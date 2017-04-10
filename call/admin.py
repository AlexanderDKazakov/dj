from django.contrib import admin
from call.models import *
from attachments.admin import AttachmentInlines
from django.contrib.admin import DateFieldListFilter

######## EXPORTING
from django.http import HttpResponse
import datetime


def export_csv(modeladmin, request, queryset):
    # FUNCTION NOT FINISHED; PROBABLY ITS WILL BE DELETED
    import csv
    from django.utils.encoding import smart_str
    # response = HttpResponse(mimetype='text/csv')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=Otchet_ot_'+str(datetime.date.today())+'.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        u'Приложение N 7 к Единым стандартам качества обслуживания сетевыми организациями потребителей услуг сетевых организаций Список изменяющих документов (введено Приказом Минэнерго России от 06.04.2015 N 217)'
    ])
    writer.writerow([
        u'Информация о качестве обслуживания потребителей_______________Филиал АО "ЛОЭСК"_____________________ услуг за 2015  год (наименование сетевой организации)'
    ])
    writer.writerow([''])
    writer.writerow([
        u'4.9. Информация по обращениям потребителей.'])
    writer.writerow([' ', ' ', ' ', ' ',
                     'Форма обращения',
                     ' ', ' ', ' ', ' ',
                     'Обращения'
                     ' ', ' ', ' ', ' ', ' ', ' ',
                     'Обращения потребителей, содержащие жалобу',
                     ' ', ' ', ' ', ' ', ' ', ' ',
                     'Обращения потребителей, содержащие заявку на оказание услуг',
                     ' ', ' ', ' ',
                     'Факт получения потребителем ответа',
                     ' ', ' ',
                     'Мероприятия по результатам обращения',
                     ' ', ' '])
    writer.writerow(
        ['N',
         'Идентификационный номер обращения',
         'Дата обращения',
         'Время обращения',
         'Очное обращение',
         'Заочное обращение посредством телефонной связи',
         'Заочное обращение посредством сети Интернет',
         'Письменное обращение посредством почтовой связи',
         'Прочее',
         'Оказание услуг по передаче электрической энергии',
         'Осуществление технологического присоединения',
         'Коммерческий учет электрической энергии',
         'Качество обслуживания потребителей',
         'Техническое обслуживание электросетевых объектов',
         'Прочее',
         'Качество услуг по передаче электрической энергии',
         'Качество электрической энергии',
         'Осуществление технологического присоединения',
         'Коммерческий учет электрической энергии',
         'Качество обслуживания потребителей',
         'Техническое обслуживание электросетевых объектов',
         'Прочее',
         'По технологическому присоединению',
         'Заключение договора на оказание услуг по передаче электроэнергии',
         'Организация коммерческого учета электроэнергии',
         'Прочее',
         'Заявителем был получен исчерпывающий ответ в установленные сроки',
         'Заявителем был получен исчерпывающий ответ с нарушением сроков',
         'Обращение оставлено без ответа',
         'Выполненные мероприятия по результатам обращения',
         'Планируемые мероприятия по результатам обращения'])
    writer.writerow(
        ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20',
         '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31'])
    calls = Call.objects.all()
    count_number = 0
    for call in sorted(calls):
        count_number += 1
        # writer.writerow(call)
        writer.writerow(
                [
                    count_number,
                    'Инд.номер',
                    # call.call_date_start.[:10],
                    ' ', ' ',
                    '+',
                    ' ', ' ', ' ',
                    '',
                ]
        )
    writer.writerow([
        smart_str(u"ID"),
        smart_str(u"call_title"),
        smart_str(u"call_aim"),
    ])
    for obj in queryset:
        writer.writerow([
            smart_str(obj.pk),
            smart_str(obj.call_title),
            smart_str(obj.call_aim),
        ])
    return response
export_csv.short_description = u"Export CSV"


def export_xls(modeladmin, request, queryset):
    import xlwt
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Otchet_ot_'+str(datetime.date.today())+'.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Отчёт от' + str(datetime.date.today()))
    # 	ws.write(0, 0, 'foobar') # row, column, value
    # TABLE HEAD
    ws.write(0, 0, u'Приложение N 7 к Единым стандартам качества обслуживания сетевыми организациями потребителей услуг сетевых организаций Список изменяющих документов (введено Приказом Минэнерго России от 06.04.2015 N 217)')
    ws.write(1, 0, u' Информация о качестве обслуживания потребителей_______________Филиал АО "ЛОЭСК"_____________________ услуг за 2015  год (наименование сетевой организации)')
    ws.write(3, 0, u'4.9. Информация по обращениям потребителей.')
    ws.write(4, 4, u'Форма обращения')
    ws.write(4, 9, u'Обращения')
    ws.write(4, 15, u'Обращения потребителей, содержащие жалобу')
    ws.write(4, 22, u'Обращения потребителей, содержащие заявку на оказание услуг')
    ws.write(4, 26, u'Факт получения потребителем ответа')
    ws.write(4, 29, u'Мероприятия по результатам обращения')
    row_list = [u'N',
                u'Идентификационный номер обращения',
                u'Дата обращения',
                u'Время обращения',
                u'Очное обращение',
                u'Заочное обращение посредством телефонной связи',
                u'Заочное обращение посредством сети Интернет',
                u'Письменное обращение посредством почтовой связи',
                u'Прочее',
                u'Оказание услуг по передаче электрической энергии',
                u'Осуществление технологического присоединения',
                u'Коммерческий учет электрической энергии',
                u'Качество обслуживания потребителей',
                u'Техническое обслуживание электросетевых объектов',
                u'Прочее',
                u'Качество услуг по передаче электрической энергии',
                u'Качество электрической энергии',
                u'Осуществление технологического присоединения',
                u'Коммерческий учет электрической энергии',
                u'Качество обслуживания потребителей',
                u'Техническое обслуживание электросетевых объектов',
                u'Прочее',
                u'По технологическому присоединению',
                u'Заключение договора на оказание услуг по передаче электроэнергии',
                u'Организация коммерческого учета электроэнергии',
                u'Прочее',
                u'Заявителем был получен исчерпывающий ответ в установленные сроки',
                u'Заявителем был получен исчерпывающий ответ с нарушением сроков',
                u'Обращение оставлено без ответа',
                u'Выполненные мероприятия по результатам обращения',
                u'Планируемые мероприятия по результатам обращения']
    # END TABLE HEAD
    # Lists for FILLING TABLE
    list_aim_call = ['Оказание услуг по передаче электрической энергии',
                     'Осуществление технологического присоединения',
                     'Коммерческий учет электрической энергии',
                     'Качество обслуживания потребителей',
                     'Техническое обслуживание электросетевых объектов',
                     'Прочее',
                     '(Жалоба) Качество услуг по передаче электрической энергии',
                     '(Жалоба) Качество электрической энергии',
                     '(Жалоба) Осуществление технологического присоединения',
                     '(Жалоба) Коммерческий учет электрической энергии',
                     '(Жалоба) Качество обслуживания потребителей',
                     '(Жалоба) Техническое обслуживание электросетевых объектов',
                     '(Жалоба) Прочее',
                     '(Заявка) По технологическому присоединению',
                     '(Заявка) Заключение договора на оказание услуг по передаче электроэнергии',
                     '(Заявка) Организация коммерческого учета электроэнергии',
                     '(Заявка) Прочее',
                     ]
    # END Lists
    # font_style = xlwt.XFStyle()
    # font_style.font.bold = True

    for index, value in enumerate(row_list):
        ws.write(5, index, str(value))
    for i in range(0, 31):
        ws.write(6, i, str(i+1))

    row_num = 6
    # columns = [
    #     (u"ID", 2000),
    #     (u"ФИО Абонента", 6000),
    #     (u"Причина звонка", 8000),
    # ]
    #
    # font_style = xlwt.XFStyle()
    # font_style.font.bold = True
    #
    # for col_num in range(len(columns)):
    #     ws.write(row_num, col_num, columns[col_num][0], font_style)
        # set column width
        # ws.col(col_num).width = columns[col_num][1]
    #
    font_style = xlwt.XFStyle()
    font_style.alignment.wrap = 1
    count_number = 0
    for obj in queryset:
        row_num += 1
        count_number += 1
        row = [
            count_number,
            row_list[1],
            str(obj.call_date)[8:10] + '/' + str(obj.call_date)[5:7] + '/' + str(obj.call_date)[:4],  # dd/mm/yyyy,
            str(obj.call_date)[11:19],
            '',
            '1',
        ]
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
        ws.write(row_num, 9+list_aim_call.index(str(obj.call_aim)), '1')
        ws.write(row_num, 26, '1') # call answer get at the time

    wb.save(response)
    return response
export_xls.short_description = u"Export XLS"


def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.utils import get_column_letter
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Otchet_ot_'+str(datetime.date.today())+'.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = 'Отчёт от ' + str(datetime.date.today())
    # TABLE HEAD
    ws.cell(row=1, column=1, value=u"Приложение N 7 к Единым стандартам качества обслуживания сетевыми организациями потребителей услуг сетевых организаций Список изменяющих документов (введено Приказом Минэнерго России от 06.04.2015 N 217)")
    ws.cell(row=2, column=1, value=u' Информация о качестве обслуживания потребителей_______________Филиал АО "ЛОЭСК"_____________________ услуг за 2015  год (наименование сетевой организации)')
    ws.cell(row=4, column=1, value=u'4.9. Информация по обращениям потребителей.')
    ws.cell(row=5, column=5, value=u'Форма обращения')
    ws.cell(row=5, column=10, value=u'Обращения')
    ws.cell(row=5, column=16, value=u'Обращения потребителей, содержащие жалобу')
    ws.cell(row=5, column=23, value=u'Обращения потребителей, содержащие заявку на оказание услуг')
    ws.cell(row=5, column=27, value=u'Факт получения потребителем ответа')
    ws.cell(row=5, column=30, value=u'Мероприятия по результатам обращения')
    row_list = [u'N',
                u'Идентификационный номер обращения',
                u'Дата обращения',
                u'Время обращения',
                u'Очное обращение',
                u'Заочное обращение посредством телефонной связи',
                u'Заочное обращение посредством сети Интернет',
                u'Письменное обращение посредством почтовой связи',
                u'Прочее',
                u'Оказание услуг по передаче электрической энергии',
                u'Осуществление технологического присоединения',
                u'Коммерческий учет электрической энергии',
                u'Качество обслуживания потребителей',
                u'Техническое обслуживание электросетевых объектов',
                u'Прочее',
                u'Качество услуг по передаче электрической энергии',
                u'Качество электрической энергии',
                u'Осуществление технологического присоединения',
                u'Коммерческий учет электрической энергии',
                u'Качество обслуживания потребителей',
                u'Техническое обслуживание электросетевых объектов',
                u'Прочее',
                u'По технологическому присоединению',
                u'Заключение договора на оказание услуг по передаче электроэнергии',
                u'Организация коммерческого учета электроэнергии',
                u'Прочее',
                u'Заявителем был получен исчерпывающий ответ в установленные сроки',
                u'Заявителем был получен исчерпывающий ответ с нарушением сроков',
                u'Обращение оставлено без ответа',
                u'Выполненные мероприятия по результатам обращения',
                u'Планируемые мероприятия по результатам обращения']
    # END TABLE HEAD
    # Lists for FILLING TABLE
    list_aim_call = ['Оказание услуг по передаче электрической энергии',
                     'Осуществление технологического присоединения',
                     'Коммерческий учет электрической энергии',
                     'Качество обслуживания потребителей',
                     'Техническое обслуживание электросетевых объектов',
                     'Прочее',
                     '(Жалоба) Качество услуг по передаче электрической энергии',
                     '(Жалоба) Качество электрической энергии',
                     '(Жалоба) Осуществление технологического присоединения',
                     '(Жалоба) Коммерческий учет электрической энергии',
                     '(Жалоба) Качество обслуживания потребителей',
                     '(Жалоба) Техническое обслуживание электросетевых объектов',
                     '(Жалоба) Прочее',
                     '(Заявка) По технологическому присоединению',
                     '(Заявка) Заключение договора на оказание услуг по передаче электроэнергии',
                     '(Заявка) Организация коммерческого учета электроэнергии',
                     '(Заявка) Прочее',
                     ]
    # END Lists
    for index, value in enumerate(row_list):
        ws.cell(row=6, column=index+1, value=str(value))
    for i in range(1, 32):
        ws.cell(row=7, column=i, value=i)

    row_num = 10
    # columns = [
    #     (u"Приложение N 7 к Единым стандартам качества обслуживания сетевыми организациями потребителей услуг сетевых организаций Список изменяющих документов (введено Приказом Минэнерго России от 06.04.2015 N 217)", 15),
    #     (u"ID", 15),
    #     (u"call_title", 7),
    #     (u"call_aim", 7),
    # ]
    #
    # for col_num in range(len(columns)):
    #     c = ws.cell(row=row_num + 1, column=col_num + 1)
    #     c.value = columns[col_num][0]
    #     c.style.font.bold = True
        # set column width
        # ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]
    count_number = 0
    for obj in queryset:
        row_num += 1
        count_number += 1
        row = [
            count_number,
            row_list[1],
            str(obj.call_date)[8:10] + '/' + str(obj.call_date)[5:7] + '/' + str(obj.call_date)[:4],  # dd/mm/yyyy
            str(obj.call_date)[11:19],
            '',
            1,
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            # c.style.alignment.wrap_text = True
        ws.cell(row=row_num, column=10 + list_aim_call.index(str(obj.call_aim)), value=1)
        ws.cell(row=row_num, column=27, value=1)  # call answer get at the time

    wb.save(response)
    return response
export_xlsx.short_description = u"Export XLSX"


class CallInline(admin.StackedInline):
    model = Comment
    extra = 2


class ProfileAdmin(admin.ModelAdmin):
    list_display = [ 'user',
                     # 'first_name',
                     'user_filial_id',
                     'user_otdel_id',
    ]


class ReasonAdmin(admin.ModelAdmin):
    list_display = [ 'id',
                     'rc_name',
                     'otdel_id_id',
    ]

class Table_NumAdmin(admin.ModelAdmin):
    list_display = ['table_name',
                    'table_reason',
                    'table_num',
                    ]


class CallAdmin(admin.ModelAdmin):
    actions = [export_xls, export_xlsx]
    list_filter = ['call_entite',
                   'call_date_start',# DateFieldListFilter,
                   'call_aim',
                   'call_document',
                   'call_user_man',
                   'call_user_man_filial',
                   'call_user_man_otdel',
                   ]
    list_display = ['id',
                    'call_title',
                    'call_aim',
                    'call_aim_detail',
                    'call_document',
                    'call_kontact',
                    'call_otvet',
                    'call_date_start',
                    'call_user_man',
                    ]
    inlines = [CallInline]

    # list_editable = ['id']
inlines = (AttachmentInlines,)
admin.site.register(Call, CallAdmin)
admin.site.register(Profile, ProfileAdmin)
# admin.site.register(Comment)
admin.site.register(Table_Num, Table_NumAdmin)
admin.site.register(Filial)
admin.site.register(legalEntity)
admin.site.register(reason_otdel, ReasonAdmin)
admin.site.register(Res)
admin.site.register(Otdel)
admin.site.register(ActOperator)

